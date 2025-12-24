import datetime
import math
import openpyxl
import pandas as pd
from pathlib import Path


def load_buyer_data(buyer_file: str, vendor_code: str) -> pd.DataFrame:
    """
    Load buyer file buyer_Axxxx.xlsx inside output_buyers folder.
    Filter to vendor + จำนวนเฉลี่ยต่อเดือน <= 6.
    """

    raw = pd.read_excel(buyer_file, header=None)

    # Row 3 = header row; row 4 onward = data
    header_row = raw.iloc[3]
    data = raw.iloc[4:].reset_index(drop=True)
    data.columns = header_row

    # Convert numeric columns
    numeric_cols = [
        "ยอดขาย", "ยอดขาย/เดือน", "min*3", "max*6",
        "สินค้าคงเหลือ", "ON_ORDER", "QTY TOTAL",
        "จำนวนเฉลี่ยต่อเดือน", "หยวน"
    ]
    for c in numeric_cols:
        if c in data.columns:
            data[c] = pd.to_numeric(data[c], errors="coerce")

    # Filter for vendor + rows with รหัสสินค้า
    df = data.copy()
    df = df[df["buyer"] == vendor_code]
    df = df[~df["รหัสสินค้า"].isna()]

    # Only <= 6
    df = df[df["จำนวนเฉลี่ยต่อเดือน"] <= 6]

    return df.reset_index(drop=True)


def fill_po(
    template_file: str,
    vendor_code: str,
    output_folder: str = "output_PO",
    date_=None,
    rate_thb_per_cny: float = 6.0,
    rows_per_page: int = 5,
):

    # ----------- Prepare paths -----------
    buyer_file = f"output_buyers/buyer_{vendor_code}.xlsx"
    buyer_file = Path(buyer_file)

    if not buyer_file.exists():
        raise FileNotFoundError(
            f"❌ File not found: {buyer_file}\n"
            "Make sure buyer_Axxxx.xlsx exists in folder output_buyers/"
        )

    # Create output folder
    Path(output_folder).mkdir(exist_ok=True)

    # PO output filename
    output_file = Path(output_folder) / f"PO_{vendor_code}.xlsx"

    # ----------- Load buyer data -----------
    df = load_buyer_data(str(buyer_file), vendor_code)

    if date_ is None:
        date_ = datetime.date.today()

    # ----------- Load template workbook -----------
    wb = openpyxl.load_workbook(template_file)
    base_ws = wb["AOTTER"]

    # ----------- Determine number of pages -----------
    total = len(df)
    pages = max(1, math.ceil(total / rows_per_page))

    sheets = [base_ws]
    for i in range(1, pages):
        ws_new = wb.copy_worksheet(base_ws)
        ws_new.title = f"AOTTER_{i+1}"
        sheets.append(ws_new)

    # ----------- Fill data -----------
    for idx, row in df.iterrows():

        page = idx // rows_per_page
        pos = idx % rows_per_page
        excel_row = 9 + pos
        ws = sheets[page]

        # Header info
        ws["H6"] = vendor_code
        ws["K6"] = date_

        # Mapping columns
        ws.cell(excel_row, 1).value = row["รหัสสินค้า"]         # A BUYER ITEM NO.
        ws.cell(excel_row, 3).value = row["รายละเอียดสินค้า"]   # C DESCRIPTION
        ws.cell(excel_row, 11).value = row["หยวน"]              # K FOB (CNY)

        # THB
        price_cny = row["หยวน"]
        ws.cell(excel_row, 12).value = None if pd.isna(price_cny) else price_cny * rate_thb_per_cny

        # USE MONTH
        ws.cell(excel_row, 14).value = row["ยอดขาย"]

        # MIN*3  /  MAX*6
        ws.cell(excel_row, 15).value = row["min*3"]
        ws.cell(excel_row, 16).value = row["max*6"]

        # STOCK GREEN / ON ORDER / TOTAL QTY
        ws.cell(excel_row, 17).value = row["สินค้าคงเหลือ"]
        ws.cell(excel_row, 18).value = row["ON_ORDER"]
        ws.cell(excel_row, 19).value = row["QTY TOTAL"]

        # จำนวนเฉลี่ยต่อเดือน
        ws.cell(excel_row, 20).value = row["จำนวนเฉลี่ยต่อเดือน"]

    # ----------- Save result -----------
    wb.save(output_file)
    print(f"✅ Exported PO file: {output_file}")
    print(f"→ Total items included: {total} (≤6-month average filter)\n")


# ------------------ RUN EXAMPLE ------------------
if __name__ == "__main__":
    fill_po(
        template_file="ตัวอย่างใบสั่งซื้อต่างประเทศ.xlsx",
        vendor_code="A0029",        # Change this anytime
        output_folder="output_PO",  # PO export folder
        rate_thb_per_cny=6.0,       # THB = 6 × CNY
        rows_per_page=5
    )
