import os
import re
from typing import List, Dict
import datetime
import math
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font
from openpyxl.drawing.image import Image as XLImage
from io import BytesIO
from PIL import Image as PILImage
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.styles import Border, Side
from openpyxl.styles import PatternFill


# ---------- CONFIG ----------
INPUT_XLSX = "update_yuan.xlsx"          # Excel report from Express
OUTPUT_FOLDER = "output_buyers"          # Folder to save per-buyer files
A0029_CATALOG_XLSX = "A0029.xlsx"
VENDOR_INFO_XLSX = "รายงานข้อมูลผู้จำหน่าย.xlsx"
# ---------------------------

THAI_MONTHS = {
    "ม.ค.": 1, "ก.พ.": 2, "มี.ค.": 3, "เม.ย.": 4,
    "พ.ค.": 5, "มิ.ย.": 6, "ก.ค.": 7, "ส.ค.": 8,
    "ก.ย.": 9, "ต.ค.": 10, "พ.ย.": 11, "ธ.ค.": 12
}


# =========================
# Utilities
# =========================
def norm_text(v) -> str:
    if v is None:
        return ""
    return str(v).replace("\xa0", " ").strip()

def norm_header(x) -> str:
    return str(x).replace("\n", " ").replace("\xa0", " ").strip() if x is not None else ""

def round_half_up(x: float) -> int:
    return int(math.floor(x + 0.5))

def load_vendor_map(path: str) -> dict:
    df = pd.read_excel(path, header=None)
    vendor_map = {}
    for _, r in df.iterrows():
        code = str(r.iloc[0]).strip() if not pd.isna(r.iloc[0]) else ""
        if not code:
            continue
        name = "" if pd.isna(r.iloc[1]) else str(r.iloc[1]).strip()
        addr = "" if pd.isna(r.iloc[2]) else str(r.iloc[2]).strip()
        vendor_map[code] = {"name": name, "address": addr}
    return vendor_map

def find_label_cell(ws, label: str, max_row=60, max_col=30):
    label = norm_text(label)
    for r in range(1, min(max_row, ws.max_row) + 1):
        for c in range(1, min(max_col, ws.max_column) + 1):
            if norm_text(ws.cell(r, c).value) == label:
                return (r, c)
    return None


# =========================
# Product field split (your logic)
# =========================
def split_product_field(s: str) -> tuple[str, str]:
    if not isinstance(s, str):
        return "", ""
    s = s.strip()
    if not s:
        return "", ""

    parts = s.split(maxsplit=1)
    if len(parts) < 2:
        return "", ""
    rest = parts[1].strip()
    if not rest:
        return "", ""

    m_th = re.search(r'[\u0E00-\u0E7F]', rest)

    if m_th:
        th_pos = m_th.start()
        if th_pos == 0:
            return "", rest

        pre_th = rest[:th_pos].strip()
        tail_th = rest[th_pos:].strip()
        tokens = pre_th.split()
        if not tokens:
            return "", tail_th

        code_idx = 0
        for i, t in enumerate(tokens):
            if re.match(r"^(No[A-Za-z0-9\-]+|[A-Z]-\d+)", t):
                code_idx = i
                break

        base_code = tokens[code_idx]
        extra_tokens = tokens[code_idx + 1:]

        tag = ""
        if extra_tokens:
            t = extra_tokens[0]
            if re.fullmatch(r"[A-Za-z]", t) or re.fullmatch(r"\([A-Za-z]\)", t):
                tag = t
                extra_tokens = extra_tokens[1:]

        code_raw = f"{base_code} {tag}".strip() if tag else base_code
        code_raw = re.sub(r'^[Nn][Oo](?=[A-Za-z0-9])', '', code_raw)

        desc_pre = " ".join(extra_tokens).strip()
        desc = (desc_pre + " " + tail_th).strip() if desc_pre else tail_th

    else:
        tokens = rest.split()
        if not tokens:
            return "", ""
        if re.search(r'[\u0E00-\u0E7F]', tokens[0]):
            return "", rest
        code_raw = tokens[0]
        desc = " ".join(tokens[1:]).strip()

    code = re.sub(r'^[Nn][Oo](?=[A-Za-z0-9])', '', code_raw)
    return code, desc


# =========================
# Date parsing (your logic)
# =========================
def thai_to_date(day: int, thai_month: str, thai_year: int) -> datetime.date:
    year = thai_year - 543
    thai_month = re.sub(r"\s+", "", thai_month)
    month = THAI_MONTHS.get(thai_month)
    if month is None:
        raise ValueError(f"Unknown Thai month token: {repr(thai_month)}")
    return datetime.date(year, month, day)

def last_day_of_month(year: int, month: int) -> int:
    if month == 12:
        next_first = datetime.date(year + 1, 1, 1)
    else:
        next_first = datetime.date(year, month + 1, 1)
    return (next_first - datetime.timedelta(days=1)).day

def add_months(dt: datetime.date, n: int) -> datetime.date:
    month = dt.month - 1 + n
    year = dt.year + month // 12
    month = month % 12 + 1
    day = min(dt.day, last_day_of_month(year, month))
    return datetime.date(year, month, day)

def calc_days_and_months(d1: datetime.date, d2: datetime.date) -> tuple[int, int]:
    if d2 < d1:
        return 0, 0
    days = (d2 - d1).days
    months = 0
    cur = d1
    while True:
        next_m = add_months(cur, 1)
        if next_m <= d2:
            months += 1
            cur = next_m
        else:
            break
    leftover_days = (d2 - cur).days
    if leftover_days >= 15:
        months += 1
    return days, months

def parse_date_range_from_header(df_raw: pd.DataFrame) -> Dict[str, object] | None:
    pattern = r"วันที่จาก\s+(\d+)\s+(\S+)\s+(\d+)\s+ถึง\s+(\d+)\s+(\S+)\s+(\d+)"
    col0 = df_raw.iloc[:, 0].astype(str)
    for val in col0:
        if "วันที่จาก" not in val:
            continue
        text = str(val).replace("\xa0", " ")
        text = re.sub(r"\s+", " ", text).strip()
        m = re.search(pattern, text)
        if not m:
            continue
        d1, m1, y1, d2, m2, y2 = m.groups()
        start_date = thai_to_date(int(d1), m1, int(y1))
        end_date = thai_to_date(int(d2), m2, int(y2))
        days, months = calc_days_and_months(start_date, end_date)
        print(">>> พบช่วงวันที่ใน header:", text)
        print("    start =", start_date, "end =", end_date,
              "days =", days, "months(กฎ 15 วัน) =", months)
        return {
            "raw_line": text.strip(),
            "start_date": start_date,
            "end_date": end_date,
            "days": days,
            "months": months,
        }
    print(">>> ไม่พบช่วงวันที่ ('วันที่จาก ... ถึง ...') ใน header")
    return None


# =========================
# Load report lines (your logic)
# =========================
def find_separator_row(df_raw: pd.DataFrame) -> int:
    print(">>> ค้นหาแถวเส้นคั่น (------)")
    for idx in range(len(df_raw)):
        row = df_raw.iloc[idx]
        for val in row:
            if isinstance(val, str) and re.search(r"-{5,}", val):
                print(f"พบแถวเส้นคั่น (------) ที่แถว: {idx}")
                return idx
    print("ไม่พบแถวเส้นคั่น (------) จะใช้ข้อมูลตั้งแต่แถวแรกแทน")
    return -1

def load_data_lines_from_excel(path: str):
    print(f"กำลังอ่านไฟล์ Excel: {path}")
    df_raw = pd.read_excel(path, header=None, dtype=str)
    print(f"จำนวนแถวทั้งหมดในไฟล์ (df_raw): {len(df_raw)}")

    date_info = parse_date_range_from_header(df_raw)

    sep_idx = find_separator_row(df_raw)
    start_idx = sep_idx + 1
    if start_idx >= len(df_raw):
        raise RuntimeError("ไม่มีข้อมูลหลังจากแถวเส้นคั่น")

    lines: List[str] = []
    for i in range(start_idx, len(df_raw)):
        row = df_raw.iloc[i]
        first_cell = row.iloc[0]
        if pd.isna(first_cell):
            continue
        base_text = str(first_cell).strip()
        if not base_text:
            continue

        yuan_val = None
        for val in reversed(row.tolist()[1:]):
            if val is None or (isinstance(val, float) and pd.isna(val)):
                continue
            s = str(val).strip()
            if not s:
                continue
            yuan_val = s
            break

        combined = f"{base_text} ||YUAN={yuan_val}" if yuan_val is not None else base_text
        lines.append(combined)

    print(f"จำนวนแถวข้อมูลหลังจากตัดแถวเส้นคั่นและแถวว่าง: {len(lines)}")
    print("\n=== DEBUG: แถวข้อมูลหลังตัดเส้นคั่น (10 แถวแรก) ===")
    for i, line in enumerate(lines[:10]):
        print(f"line[{i}] = {repr(line)}")
    print("=================================================\n")
    return lines, date_info

def is_header_or_separator(line: str) -> bool:
    s = line.strip()
    if not s:
        return True
    if "BUYER" in s:
        return True
    if re.fullmatch(r"-{10,}", s):
        return True
    if re.search(r"-{10,}", s):
        return True
    return False

def parse_line_to_fields(line: str) -> Dict[str, str] | None:
    yuan_from_tail = ""
    m_y = re.search(r"\|\|YUAN=(\S+)", line)
    if m_y:
        yuan_from_tail = m_y.group(1)
        line = line[:m_y.start()].rstrip()

    m = re.match(r"\s*([0-9A-Za-z]{5})\b(.*)", line)
    if not m:
        return None
    buyer = m.group(1)
    rest = m.group(2)

    tokens_all = re.split(r"\s+", rest.strip())
    if len(tokens_all) < 2:
        return None

    barcode = ""
    idx = 0
    if re.fullmatch(r"\d+", tokens_all[idx]):
        barcode = tokens_all[idx]
        idx += 1
    if idx < len(tokens_all) and tokens_all[idx] == "}":
        idx += 1

    tokens_all = tokens_all[idx:]
    if len(tokens_all) < 2:
        return None

    def is_num_token(s: str) -> bool:
        s = s.replace(",", "").replace('"', "").strip()
        return bool(re.fullmatch(r"\d+(\.\d+)?", s))

    on_order = "0"
    last_tok = tokens_all[-1]
    if is_num_token(last_tok):
        on_order = last_tok.replace('"', "")
        tokens_all = tokens_all[:-1]

    numeric = []
    non_numeric_rev = []
    for tok in reversed(tokens_all):
        if is_num_token(tok):
            numeric.append(tok.replace('"', ""))
        else:
            non_numeric_rev.append(tok)
    numeric = list(reversed(numeric))
    non_numeric = list(reversed(non_numeric_rev))

    if not numeric:
        return None

    if len(numeric) > 5:
        numeric = numeric[-5:]
    while len(numeric) < 4:
        numeric.insert(0, "0.00")

    if len(numeric) == 4:
        qty0, val0, qty_balance, val_balance = numeric
        qty_sale = "0.00"
    else:
        qty0, val0, qty_sale, qty_balance, val_balance = numeric[-5:]

    product_and_desc = " ".join(non_numeric).strip()
    yuan = yuan_from_tail

    return {
        "buyer": buyer,
        "barcode": barcode,
        "สินค้า": product_and_desc,
        "ยกยอดมา": qty0,
        "มูลค่า": val0,
        "ยอดขาย": qty_sale,
        "สินค้าคงเหลือ": qty_balance,
        "มูลค่า2": val_balance,
        "หยวน": yuan,
        "ON_ORDER": on_order,
    }

def clean_yuan_value(val):
    if pd.isna(val):
        return np.nan
    s = str(val).strip()
    if not s:
        return np.nan
    if re.fullmatch(r"\d+(\.\d+)?", s):
        return float(s)
    return np.nan


# =========================
# Buyer file -> PO data
# =========================
def load_buyer_data_for_po(buyer_xlsx_path: str, vendor_code: str) -> pd.DataFrame:
    raw = pd.read_excel(buyer_xlsx_path, header=None)
    header_row = raw.iloc[3]
    data = raw.iloc[4:].reset_index(drop=True)
    data.columns = header_row

    numeric_cols = [
        "ยอดขาย", "ยอดขาย/เดือน", "min*3", "max*6",
        "สินค้าคงเหลือ", "ON_ORDER", "QTY TOTAL", "จำนวนเฉลี่ยต่อเดือน", "หยวน",
    ]
    for col in numeric_cols:
        if col in data.columns:
            data[col] = pd.to_numeric(data[col], errors="coerce")

    df = data.copy()
    df = df[df["buyer"] == vendor_code]
    df = df[~df["รหัสสินค้า"].isna()]
    df = df[(df["จำนวนเฉลี่ยต่อเดือน"] <= 7) & (df["ยอดขาย"] != 0)]
    return df.reset_index(drop=True)


# =========================
# Catalog map + images
# =========================
def build_catalog_map_from_A0029(a0029_path: str):
    wb = openpyxl.load_workbook(a0029_path)
    ws = wb.active

    header = [norm_text(c.value) for c in ws[1]]

    def col(name):
        return header.index(name) + 1

    img_at = {}
    for img in ws._images:
        try:
            r = img.anchor._from.row + 1
            c = img.anchor._from.col + 1
            img_bytes = img._data()
            img_at[(r, c)] = img_bytes
        except Exception as e:
            print("⚠️ Cannot read an image in A0029:", e)

    pic_col = col("GOODS PICTURE")

    catalog = {}
    for r in range(2, ws.max_row + 1):
        item_no = ws.cell(r, col("BUYER ITEM NO.")).value
        if not item_no:
            continue
        item_no = str(item_no).strip()

        catalog[item_no] = {
            "goods_desc": ws.cell(r, col("GOODS DESCRIPTION")).value,
            "brand": ws.cell(r, col("BRAND")).value,
            "material": ws.cell(r, col("MATERIAL")).value,
            "weight": ws.cell(r, col("Weight")).value,
            "qty_per_carton": ws.cell(r, col("QTY PER CARTON")).value,
            "img_bytes": img_at.get((r, pic_col)),
        }

    print(f">>> Catalog loaded: {len(catalog)} items, images found: {len(img_at)}")
    return catalog


# =========================
# Image centering in cell (your functions)
# =========================
def _excel_colwidth_to_pixels(width):
    if width is None:
        width = 8.43
    return int(width * 7 + 5)

def _excel_rowheight_to_pixels(height_pts):
    if height_pts is None:
        height_pts = 15
    return int(height_pts * 96 / 72)

def _get_cell_rect_pixels(ws, col_letter, row_num):
    col_w = _excel_colwidth_to_pixels(ws.column_dimensions[col_letter].width)
    row_h = _excel_rowheight_to_pixels(ws.row_dimensions[row_num].height)

    for mr in ws.merged_cells.ranges:
        if mr.min_col <= column_index_from_string(col_letter) <= mr.max_col and mr.min_row <= row_num <= mr.max_row:
            total_w = 0
            for c in range(mr.min_col, mr.max_col + 1):
                letter = openpyxl.utils.get_column_letter(c)
                total_w += _excel_colwidth_to_pixels(ws.column_dimensions[letter].width)
            total_h = 0
            for r in range(mr.min_row, mr.max_row + 1):
                total_h += _excel_rowheight_to_pixels(ws.row_dimensions[r].height)
            return total_w, total_h

    return col_w, row_h

def add_image_to_cell(ws, cell_addr: str, img_bytes: bytes, max_w=90, max_h=90):
    if not img_bytes:
        return

    pil = PILImage.open(BytesIO(img_bytes)).convert("RGBA")
    w, h = pil.size
    scale = min(max_w / w, max_h / h, 1.0)
    new_w = max(1, int(w * scale))
    new_h = max(1, int(h * scale))
    pil = pil.resize((new_w, new_h))

    bio = BytesIO()
    pil.save(bio, format="PNG")
    bio.seek(0)

    img = XLImage(bio)
    img.width = new_w
    img.height = new_h

    if not hasattr(ws, "_img_buffers"):
        ws._img_buffers = []
    ws._img_buffers.append(bio)

    col_letter, row_num = coordinate_from_string(cell_addr)
    row_num = int(row_num)
    col_idx0 = column_index_from_string(col_letter) - 1
    row_idx0 = row_num - 1

    cell_w_px, cell_h_px = _get_cell_rect_pixels(ws, col_letter, row_num)

    x_off_px = max(0, int((cell_w_px - new_w) / 2))
    y_off_px = max(0, int((cell_h_px - new_h) / 2))

    marker = AnchorMarker(col=col_idx0, colOff=pixels_to_EMU(x_off_px),
                          row=row_idx0, rowOff=pixels_to_EMU(y_off_px))

    img.anchor = OneCellAnchor(
        _from=marker,
        ext=XDRPositiveSize2D(pixels_to_EMU(new_w), pixels_to_EMU(new_h))
    )
    ws.add_image(img)


# =========================
# PO column map
# =========================
def get_po_col_map(ws, header_row=8):
    col_map = {}
    for c in range(1, ws.max_column + 1):
        v = norm_header(ws.cell(header_row, c).value)
        if v:
            col_map[v] = c
    return col_map


# =========================
# NEW: expand single-page rows
# =========================
def expand_po_rows(ws, item_start_row: int, base_item_rows: int, n_items: int, remark_row: int) -> int:
    """
    items start at row 9
    base_item_rows = 5
    remark_row = 14 (when 5 items)
    If n_items > 5, insert extra rows BEFORE remark_row so remark shifts down.
    Return new remark_row.
    """
    extra = max(0, n_items - base_item_rows)
    if extra <= 0:
        return remark_row
    ws.insert_rows(remark_row, amount=extra)
    return remark_row + extra

def set_uniform_item_row_heights(ws, item_start_row: int, n_items: int, template_item_row: int):
    """
    Make all item rows have the same height as the template item row.
    """
    base_h = ws.row_dimensions[template_item_row].height
    if base_h is None:
        base_h = 15  # Excel default fallback

    for r in range(item_start_row, item_start_row + n_items):
        ws.row_dimensions[r].height = base_h

def set_item_font(ws, item_start_row: int, n_items: int, font_name="Arial", font_size=18):
    """
    Set font for all item cells (entire row range) to Arial 18.
    """
    f = Font(name=font_name, size=font_size)


def set_item_font(ws, item_start_row: int, n_items: int, font_name="Arial", font_size=18):
    """
    Set font for all item cells (entire row range) to Arial 18.
    """
    f = Font(name=font_name, size=font_size)

    for r in range(item_start_row, item_start_row + n_items):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).font = f



from openpyxl.styles import Alignment

def center_item_cells(ws, item_start_row: int, n_items: int):
    align = Alignment(vertical="center", horizontal="center")
    for r in range(item_start_row, item_start_row + n_items):
        for c in range(1, ws.max_column + 1):
            ws.cell(r, c).alignment = align

def add_signature_under_last_item(
    ws,
    last_item_row: int,
    remark_row: int,
    sig_path: str = "footer_signatures.png",
    gap_rows: int = 2,
    min_gap_below_remark: int = 2,
    anchor_col: str = "A",
):
    """
    Put signature underneath the last product row.
    Also ensure it is not above the remark block (in case remark_row is lower).
    """
    if not os.path.exists(sig_path):
        return

    # candidate row directly under last item
    sig_row = last_item_row + gap_rows

    # make sure it's below remark too (so it never overlaps)
    sig_row = max(sig_row, remark_row + min_gap_below_remark)

    sig = XLImage(sig_path)
    sig.width = 1000
    sig.height = 750
    ws.add_image(sig, f"{anchor_col}{sig_row}")
def apply_item_borders(
    ws,
    item_start_row: int,
    n_items: int,
    green_col: int,
    total_green_col: int,
    last_col: int = 21,   # ⬅ column U
):
    """
    - Thin border for all item cells up to column U
    - Thick border wrapping GREEN + TOTAL QTY GREEN columns together
    """

    thin = Side(style="thin")
    thick = Side(style="thick")

    thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    first_row = item_start_row
    last_row = item_start_row + n_items - 1

    # ---- THIN BORDER for all item cells (A → U only) ----
    for r in range(first_row, last_row + 1):
        for c in range(1, last_col + 1):
            ws.cell(r, c).border = thin_border

    # ---- THICK BORDER around GREEN + TOTAL QTY GREEN ----
    for r in range(first_row, last_row + 1):

        top = thick if r == first_row else thin
        bottom = thick if r == last_row else thin

        # LEFT edge (GREEN)
        ws.cell(r, green_col).border = Border(
            left=thick, right=thin, top=top, bottom=bottom
        )

        # MIDDLE columns (if any)
        for c in range(green_col + 1, total_green_col):
            ws.cell(r, c).border = Border(
                left=thin, right=thin, top=top, bottom=bottom
            )

        # RIGHT edge (TOTAL QTY GREEN)
        ws.cell(r, total_green_col).border = Border(
            left=thin, right=thick, top=top, bottom=bottom
        )

def apply_column_fills(ws, item_start_row: int, n_items: int):
    """
    Apply fixed background colors to specific columns for item rows.
    """

    fills = {
        "J": PatternFill("solid", start_color="DDEBF7", end_color="DDEBF7"),  # light blue
        "N": PatternFill("solid", start_color="E2F0D9", end_color="E2F0D9"),  # light green
        "O": PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC"),  # light yellow
        "P": PatternFill("solid", start_color="F8CBAD", end_color="F8CBAD"),  # light red
        "Q": PatternFill("solid", start_color="E2F0D9", end_color="E2F0D9"),  # light green
    }

    for r in range(item_start_row, item_start_row + n_items):
        for col_letter, fill in fills.items():
            ws[f"{col_letter}{r}"].fill = fill


def highlight_item_rows(
    ws,
    item_start_row: int,
    n_items: int,
    last_col: int = 21,   # Column U
    fill_color: str = "FFFFFF",  # default = white (change below)
):
    """
    Apply background fill to item rows from A to U.
    fill_color = hex RGB (no #)
    """
    fill = PatternFill(
        fill_type="solid",
        start_color=fill_color,
        end_color=fill_color,
    )

    for r in range(item_start_row, item_start_row + n_items):
        for c in range(1, last_col + 1):
            ws.cell(r, c).fill = fill

def apply_currency_formats_dynamic(ws, item_start_row: int, n_items: int, po_cols: dict, remark_row: int):
    fmt_cny = '"¥" #,##0.00'
    fmt_thb = '"฿" #,##0.00'

    # item rows
    for r in range(item_start_row, item_start_row + n_items):
        if "FOB PRICE (CNY)" in po_cols:
            ws.cell(r, po_cols["FOB PRICE (CNY)"]).number_format = fmt_cny
        if "AMOUNT (CNY)" in po_cols:
            ws.cell(r, po_cols["AMOUNT (CNY)"]).number_format = fmt_cny
        if "THB" in po_cols:
            ws.cell(r, po_cols["THB"]).number_format = fmt_thb

    # remark totals row (same columns)
    if "AMOUNT (CNY)" in po_cols:
        ws.cell(remark_row, po_cols["AMOUNT (CNY)"]).number_format = fmt_cny
    if "THB" in po_cols:
        ws.cell(remark_row, po_cols["THB"]).number_format = fmt_thb


# =========================
# Generate PO (SINGLE SHEET, expands rows)
# =========================
def generate_po_from_template(
    template_path: str,
    vendor_code: str,
    po_date: datetime.date | None = None,
    rate_thb_per_cny: float = 6.0,
    po_output_folder: str | None = "output_PO",
    output_stream: BytesIO | None = None,
    a0029_catalog_path: str | None = None,
):
    # Fixed layout positions you confirmed
    ITEM_START_ROW = 9
    BASE_ITEM_ROWS = 5
    REMARK_ROW_BASE = 14     # remark row when 5 items
    HEADER_ROW = 8

    if po_date is None:
        po_date = datetime.date.today()

    vendor_map = load_vendor_map(VENDOR_INFO_XLSX) if os.path.exists(VENDOR_INFO_XLSX) else {}
    supplier_name = vendor_map.get(vendor_code, {}).get("name", "")
    supplier_addr = vendor_map.get(vendor_code, {}).get("address", "")

    buyer_path = os.path.join(OUTPUT_FOLDER, f"buyer_{vendor_code}.xlsx")
    if not os.path.exists(buyer_path):
        print(f"File not found: {buyer_path}")
        return

    df = load_buyer_data_for_po(buyer_path, vendor_code)

    # Filter: remaining_old <= 6 (your original rule)
    def _to_float(v):
        try:
            return float(v) if not pd.isna(v) else 0.0
        except:
            return 0.0

    monthly_usage_num_series = df["ยอดขาย/เดือน"].apply(_to_float)
    qty_total_num_series = df["QTY TOTAL"].apply(_to_float)
    remaining_old_series = np.where(
        monthly_usage_num_series > 0,
        np.floor((qty_total_num_series / monthly_usage_num_series) + 0.5),
        0.0
    )
    df = df[remaining_old_series <= 6].reset_index(drop=True)

    if len(df) == 0:
        print("No items to generate PO (after filtering).")
        return

    # Catalog
    catalog_map = {}
    if a0029_catalog_path and os.path.exists(a0029_catalog_path):
        catalog_map = build_catalog_map_from_A0029(a0029_catalog_path)

    # Output path
    if output_stream is None:
        if po_output_folder is None:
            po_output_folder = "output_PO"
        os.makedirs(po_output_folder, exist_ok=True)
        output_path = os.path.join(po_output_folder, f"PO_{vendor_code}.xlsx")
    else:
        output_path = None

    # Load workbook and use ONLY ONE sheet template: "page1"
    wb = openpyxl.load_workbook(template_path)

    if "page1" not in wb.sheetnames:
        raise RuntimeError("Your template must have a sheet named 'page1' (single layout).")

    template_ws = wb["page1"]
    ws = wb.copy_worksheet(template_ws)
    ws.title = "PO"

    # Remove all other sheets
    for sh in list(wb.worksheets):
        if sh.title != "PO":
            wb.remove(sh)

    # Expand rows so PO can hold >5 items
    remark_row = expand_po_rows(
        ws,
        item_start_row=ITEM_START_ROW,
        base_item_rows=BASE_ITEM_ROWS,
        n_items=len(df),
        remark_row=REMARK_ROW_BASE
    )

    # 🔒 Force uniform row height for all item rows
    set_uniform_item_row_heights(
        ws,
        item_start_row=ITEM_START_ROW,
        n_items=len(df),
        template_item_row=ITEM_START_ROW  # row 9 is the template row
    )

    # 🔤 Force Arial 18 font for all item rows
    set_item_font(
        ws,
        item_start_row=ITEM_START_ROW,
        n_items=len(df),
        font_name="Arial",
        font_size=18
    )

    # Column map
    po_cols = get_po_col_map(ws, header_row=HEADER_ROW)

    # Header fields
    ws.column_dimensions["B"].width = 16
    ws["H6"] = vendor_code
    ws["H6"].font = Font(color="FF0000", bold=True, size=18)
    ws["K6"] = po_date

    # Logo ALWAYS on first (only) page
    if os.path.exists("logo.png"):
        logo_img = XLImage("logo.png")
        logo_img.width = 250
        logo_img.height = 120
        ws.add_image(logo_img, "A1")

    # Supplier/Address
    pos = find_label_cell(ws, "SUPPLIER")
    if pos:
        r, c = pos
        ws.cell(r, c + 1).value = supplier_name

    pos = find_label_cell(ws, "ADDRESS")
    if pos:
        r, c = pos
        ws.cell(r, c + 1).value = supplier_addr

    # Signature (keep fixed anchor; if you want it under remark, tell me)
    if os.path.exists("footer_signatures.png"):
        sig = XLImage("footer_signatures.png")
        sig.width = 1000
        sig.height = 750

    # Totals accumulators (for remark row)
    sum_cartons = 0.0
    sum_green = 0.0
    sum_total_qty_green = 0.0
    sum_amount_cny = 0.0

    # Write items
    TARGET_MONTHS = 7

    for i, row in df.iterrows():
        line = ITEM_START_ROW + i

        buyer_item = str(row["รหัสสินค้า"]).strip()
        cat = catalog_map.get(buyer_item, {})

        # numeric safe
        try:
            monthly_usage_num = float(row.get("ยอดขาย/เดือน", 0)) if not pd.isna(row.get("ยอดขาย/เดือน", 0)) else 0.0
        except:
            monthly_usage_num = 0.0

        try:
            old_total_qty_num = float(row.get("QTY TOTAL", 0)) if not pd.isna(row.get("QTY TOTAL", 0)) else 0.0
        except:
            old_total_qty_num = 0.0

        qty_per_carton = cat.get("qty_per_carton", "")
        try:
            qty_per_carton_num = float(qty_per_carton) if qty_per_carton not in [None, ""] else 0.0
        except:
            qty_per_carton_num = 0.0

        if monthly_usage_num > 0:
            remaining_old = round_half_up(old_total_qty_num / monthly_usage_num)
        else:
            remaining_old = 0

        gap_months = max(0, TARGET_MONTHS - remaining_old)
        units_to_order = gap_months * monthly_usage_num

        if qty_per_carton_num > 0:
            cartons = round_half_up(units_to_order / qty_per_carton_num)
        else:
            cartons = 0

        green_qty = cartons * qty_per_carton_num

        new_total_qty = old_total_qty_num + green_qty
        new_remaining_months = round_half_up(new_total_qty / monthly_usage_num) if monthly_usage_num > 0 else 0

        comment_text = None
        if cartons == 0:
            comment_text = "too small order" if qty_per_carton_num > 0 else "not enough data"

        # ---- WRITE CELLS ----
        ws.cell(line, 1).value = buyer_item

        if cat.get("img_bytes"):
            add_image_to_cell(ws, f"B{line}", cat["img_bytes"], max_w=90, max_h=90)

        ws.cell(line, 3).value = cat.get("goods_desc", "")
        ws.cell(line, 4).value = cat.get("brand", "")
        ws.cell(line, 5).value = cat.get("material", "")
        ws.cell(line, 6).value = cat.get("weight", "")
        ws.cell(line, 7).value = qty_per_carton_num if qty_per_carton_num > 0 else ""

        # Key numeric columns (must exist in header map)
        if "CARTONS" in po_cols:
            ws.cell(line, po_cols["CARTONS"]).value = cartons
        if "GREEN" in po_cols:
            ws.cell(line, po_cols["GREEN"]).value = green_qty
        if "TOTAL QTY GREEN" in po_cols:
            ws.cell(line, po_cols["TOTAL QTY GREEN"]).value = green_qty

        yuan = row.get("หยวน", "")
        if "FOB PRICE (CNY)" in po_cols:
            ws.cell(line, po_cols["FOB PRICE (CNY)"]).value = yuan

        try:
            yuan_num = float(yuan) if not pd.isna(yuan) else None
        except:
            yuan_num = None

        if "THB" in po_cols:
            ws.cell(line, po_cols["THB"]).value = None if yuan_num is None else yuan_num * float(rate_thb_per_cny)

        amount_cny = None if yuan_num is None else yuan_num * float(green_qty)
        if "AMOUNT (CNY)" in po_cols:
            ws.cell(line, po_cols["AMOUNT (CNY)"]).value = amount_cny

        if "USE MONTH" in po_cols:
            ws.cell(line, po_cols["USE MONTH"]).value = monthly_usage_num
        if "MIN*3" in po_cols:
            ws.cell(line, po_cols["MIN*3"]).value = row.get("min*3", "")
        if "MAX*6" in po_cols:
            ws.cell(line, po_cols["MAX*6"]).value = row.get("max*6", "")
        if "STOCK GREEN" in po_cols:
            ws.cell(line, po_cols["STOCK GREEN"]).value = row.get("สินค้าคงเหลือ", "")
        if "ON ORDER" in po_cols:
            ws.cell(line, po_cols["ON ORDER"]).value = row.get("ON_ORDER", "")
        if "TOTAL QTY" in po_cols:
            ws.cell(line, po_cols["TOTAL QTY"]).value = old_total_qty_num

        if "จน./USE MONTH" in po_cols:
            ws.cell(line, po_cols["จน./USE MONTH"]).value = comment_text if comment_text else new_remaining_months
        if "คงเหลือ (จน./USE MONTH เดิม)" in po_cols:
            ws.cell(line, po_cols["คงเหลือ (จน./USE MONTH เดิม)"]).value = remaining_old

        # ---- ACCUMULATE TOTALS FOR REMARK ROW ----
        sum_cartons += float(cartons)
        sum_green += float(green_qty)
        sum_total_qty_green += float(green_qty)
        if amount_cny is not None:
            sum_amount_cny += float(amount_cny)

    apply_currency_formats_dynamic(
        ws,
        item_start_row=ITEM_START_ROW,
        n_items=len(df),
        po_cols=po_cols,
        remark_row=remark_row
    )

    apply_column_fills(
        ws,
        item_start_row=ITEM_START_ROW,
        n_items=len(df),
    )

    apply_item_borders(
        ws,
        item_start_row=ITEM_START_ROW,
        n_items=len(df),
        green_col=po_cols["GREEN"],
        total_green_col=po_cols["TOTAL QTY GREEN"],
    )

    last_item_row = ITEM_START_ROW + len(df) - 1

    add_signature_under_last_item(
        ws,
        last_item_row=last_item_row,
        remark_row=remark_row,
        sig_path="footer_signatures.png",
        gap_rows=4,  # signature 2 rows below last product
        min_gap_below_remark=4  # and at least 2 rows below remark
    )

    # ✅ Force Arial 18 after writing all item values
    set_item_font(ws, ITEM_START_ROW, len(df), "Arial", 18)
    center_item_cells(ws, ITEM_START_ROW, len(df))

    # Write totals into REMARK row (dynamic)
    if "CARTONS" in po_cols:
        ws.cell(remark_row, po_cols["CARTONS"]).value = sum_cartons
    if "GREEN" in po_cols:
        ws.cell(remark_row, po_cols["GREEN"]).value = sum_green
    if "TOTAL QTY GREEN" in po_cols:
        ws.cell(remark_row, po_cols["TOTAL QTY GREEN"]).value = sum_total_qty_green
    if "AMOUNT (CNY)" in po_cols:
        ws.cell(remark_row, po_cols["AMOUNT (CNY)"]).value = sum_amount_cny

    # Save
    if output_stream is not None:
        wb.save(output_stream)
        output_stream.seek(0)
        print(f"✔ PO created in memory for {vendor_code}")
        return
    else:
        wb.save(output_path)
        print(f"✔ PO created: {output_path}")
        return output_path


# =========================
# MAIN: split buyer files (same as your logic)
# =========================
def main():
    print(">>> เริ่มโปรแกรม split_buyers_local (PARSE FROM TEXT MODE)")
    lines, date_info = load_data_lines_from_excel(INPUT_XLSX)

    parsed_rows: List[Dict[str, str]] = []
    for line in lines:
        if is_header_or_separator(line):
            continue
        row = parse_line_to_fields(line)
        if row is not None:
            parsed_rows.append(row)

    print(f"จำนวนแถวที่ parse ได้สำเร็จ: {len(parsed_rows)}")
    if not parsed_rows:
        print("!!! ยัง parse ข้อมูลไม่ได้เลย ตรวจสอบรูปแบบไฟล์อีกครั้ง")
        return

    df = pd.DataFrame(parsed_rows)

    if "หยวน" in df.columns:
        df["หยวน"] = df["หยวน"].apply(clean_yuan_value)

    df[["รหัสสินค้า", "รายละเอียดสินค้า"]] = df["สินค้า"].apply(lambda x: pd.Series(split_product_field(x)))
    df.drop(columns=["สินค้า"], inplace=True)

    os.makedirs(OUTPUT_FOLDER, exist_ok=True)
    buyers = sorted(df["buyer"].unique())
    print(f"พบผู้ซื้อทั้งหมด: {len(buyers)} ราย")

    for buyer_code in buyers:
        df_buyer = df[df["buyer"] == buyer_code].copy()

        numeric_cols = ["ยกยอดมา", "มูลค่า", "ยอดขาย", "สินค้าคงเหลือ", "มูลค่า2", "ON_ORDER"]
        for col in numeric_cols:
            if col in df_buyer.columns:
                df_buyer[col] = (
                    df_buyer[col].astype(str)
                    .str.replace(",", "", regex=False)
                    .str.replace('"', "", regex=False)
                    .str.strip()
                    .replace("", "0")
                    .astype(float)
                )

        months = date_info["months"] if (date_info and date_info["months"] > 0) else 1

        sale_qty = np.floor(df_buyer["ยอดขาย"] + 0.5)
        df_buyer["ยอดขาย"] = sale_qty

        sale_per_month = sale_qty / months
        df_buyer["ยอดขาย/เดือน"] = np.floor(sale_per_month + 0.5)

        df_buyer["min*3"] = df_buyer["ยอดขาย/เดือน"] * 3
        df_buyer["max*6"] = df_buyer["ยอดขาย/เดือน"] * 6

        df_buyer["QTY TOTAL"] = df_buyer["สินค้าคงเหลือ"] + df_buyer["ON_ORDER"]

        avg_ratio = np.where(
            df_buyer["ยอดขาย/เดือน"] > 0,
            df_buyer["QTY TOTAL"] / df_buyer["ยอดขาย/เดือน"],
            0.0
        )
        df_buyer["จำนวนเฉลี่ยต่อเดือน"] = np.floor(avg_ratio + 0.5)

        df_buyer = df_buyer[
            [
                "buyer", "barcode", "รหัสสินค้า", "รายละเอียดสินค้า",
                "ยกยอดมา", "มูลค่า", "ยอดขาย",
                "ยอดขาย/เดือน", "min*3", "max*6",
                "สินค้าคงเหลือ", "มูลค่า2", "ON_ORDER",
                "QTY TOTAL", "จำนวนเฉลี่ยต่อเดือน",
                "หยวน",
            ]
        ]

        safe_buyer = re.sub(r"[^A-Za-z0-9ก-ฮะ-์_()-]+", "_", str(buyer_code).strip())[:60]
        filename = f"buyer_{safe_buyer}.xlsx"
        out_path = os.path.join(OUTPUT_FOLDER, filename)

        with pd.ExcelWriter(out_path, engine="xlsxwriter") as writer:
            if date_info is not None:
                info_df = pd.DataFrame([{
                    "ช่วงวันที่": date_info["raw_line"],
                    "จำนวนวัน": date_info["days"],
                    "เดือน": date_info["months"],
                }])
                info_df.to_excel(writer, sheet_name="Sheet1", index=False, startrow=0)
                data_start_row = len(info_df) + 2
            else:
                data_start_row = 0

            df_buyer.to_excel(writer, sheet_name="Sheet1", index=False, startrow=data_start_row)

            workbook = writer.book
            worksheet = writer.sheets["Sheet1"]

            fmt_blue = workbook.add_format({"bg_color": "#DCE6F1"})
            fmt_yellow = workbook.add_format({"bg_color": "#FFF2CC"})
            fmt_red = workbook.add_format({"bg_color": "#F8CBAD"})

            header_list = df_buyer.columns.tolist()
            col_idx_sales_month = header_list.index("ยอดขาย/เดือน")
            col_idx_min3 = header_list.index("min*3")
            col_idx_max6 = header_list.index("max*6")
            col_idx_qtytotal = header_list.index("QTY TOTAL")

            start_row = data_start_row + 1
            end_row = data_start_row + len(df_buyer)

            worksheet.conditional_format(start_row, col_idx_sales_month, end_row, col_idx_sales_month,
                                         {"type": "no_errors", "format": fmt_blue})
            worksheet.conditional_format(start_row, col_idx_min3, end_row, col_idx_min3,
                                         {"type": "no_errors", "format": fmt_yellow})
            worksheet.conditional_format(start_row, col_idx_max6, end_row, col_idx_max6,
                                         {"type": "no_errors", "format": fmt_red})
            worksheet.conditional_format(start_row, col_idx_qtytotal, end_row, col_idx_qtytotal,
                                         {"type": "no_errors", "format": fmt_red})

        print(f"✓ Saved: {out_path}")

    print(f"\nเสร็จแล้ว! ไฟล์ทั้งหมดถูกบันทึกในโฟลเดอร์: {os.path.abspath(OUTPUT_FOLDER)}")


if __name__ == "__main__":
    main()

    print("\n--- สร้างไฟล์ใบสั่งซื้อต่างประเทศ (PO) จาก buyer_xxxx.xlsx ---")
    ans = input("ต้องการสร้าง PO เลยหรือไม่? (y/n): ").strip().lower()

    if ans == "y":
        vendor = input("กรุณาระบุรหัส Vendor เช่น A0029: ").strip()
        if vendor:
            date_str = input("ระบุวันที่ PO (รูปแบบ YYYY-MM-DD, ว่าง = วันนี้): ").strip()
            if date_str:
                try:
                    po_date = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
                except ValueError:
                    print("รูปแบบวันที่ไม่ถูกต้อง ใช้วันที่วันนี้แทน")
                    po_date = datetime.date.today()
            else:
                po_date = datetime.date.today()

            template_file = "ตัวอย่างใบสั่งซื้อต่างประเทศ.xlsx"
            a0029_catalog_file = "A0029.xlsx"

            generate_po_from_template(
                template_path=template_file,
                vendor_code=vendor,
                po_date=po_date,
                rate_thb_per_cny=6.0,
                po_output_folder="output_PO",
                a0029_catalog_path=a0029_catalog_file,
            )
        else:
            print("ไม่ได้ระบุ Vendor code, ข้ามการสร้าง PO")
    else:
        print("ข้ามการสร้าง PO")
