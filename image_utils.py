import re
from io import BytesIO
from typing import Dict, Optional

import openpyxl
from openpyxl.drawing.image import Image as XLImage

# --- helpers ---
SPACE_CHARS = r"\u0020\u00A0\u1680\u2000-\u200A\u202F\u205F\u3000"

def norm_text(x) -> str:
    if x is None:
        return ""
    s = str(x)
    s = re.sub(f"[{SPACE_CHARS}]+", " ", s)
    return s.strip()

def extract_code_from_goods_description(desc: str) -> str:
    """
    From A0029 "GOODS DESCRIPTION", try to extract product code.
    Example:
      "MC-333F สต็อปวาล์ว..."  -> "MC-333F"
      "SL-L730C Shower rain(" -> "SL-L730C"
    """
    s = norm_text(desc)
    if not s:
        return ""
    # take first token up to space
    first = s.split()[0]
    return first

def build_image_map_from_A0029(a0029_path: str) -> Dict[str, bytes]:
    """
    Reads A0029 catalog.
    Header row contains:
      BUYER ITEM NO. | GOODS PICTURE | GOODS DESCRIPTION | BRAND | MATERIAL | Weight | QTY PER CARTON | UNIT PRICE (FOB)

    We use:
      - embedded images anchored in the "GOODS PICTURE" column
      - product code from the FIRST TOKEN of "GOODS DESCRIPTION" (e.g. 'MC-333F ...' -> 'MC-333F')
    """
    wb = openpyxl.load_workbook(a0029_path)
    ws = wb.active

    # locate header row
    header_row = 1
    for r in range(1, min(50, ws.max_row) + 1):
        row_vals = [norm_text(ws.cell(r, c).value) for c in range(1, min(30, ws.max_column) + 1)]
        if "GOODS PICTURE" in row_vals and "GOODS DESCRIPTION" in row_vals:
            header_row = r
            break

    # detect needed columns
    col_picture = None
    col_desc = None
    for c in range(1, ws.max_column + 1):
        v = norm_text(ws.cell(header_row, c).value)
        if v == "GOODS PICTURE":
            col_picture = c
        elif v == "GOODS DESCRIPTION":
            col_desc = c

    # fallbacks if not found
    if col_picture is None:
        col_picture = 2
    if col_desc is None:
        col_desc = 3

    img_map: Dict[str, bytes] = {}

    for img in ws._images:
        r = img.anchor._from.row + 1
        c = img.anchor._from.col + 1

        # only accept images anchored in GOODS PICTURE column
        if c != col_picture:
            continue
        if r <= header_row:
            continue

        desc = norm_text(ws.cell(r, col_desc).value)
        code = extract_code_from_goods_description(desc)
        if not code:
            continue

        try:
            img_bytes = img._data()
        except Exception:
            continue

        img_map[code] = img_bytes

    return img_map

def find_header_row(ws, max_scan=50) -> int:
    targets = {"buyer", "barcode", "รหัสสินค้า", "รายละเอียดสินค้า"}
    for r in range(1, min(max_scan, ws.max_row) + 1):
        vals = [norm_text(ws.cell(r, c).value) for c in range(1, min(40, ws.max_column) + 1)]
        if any(v in targets for v in vals):
            return r
    return 1

def find_col_by_header(ws, header_row: int, name: str) -> Optional[int]:
    for c in range(1, ws.max_column + 1):
        if norm_text(ws.cell(header_row, c).value) == name:
            return c
    return None

def add_picture_column_to_buyer_file(
    buyer_xlsx_path: str,
    img_map: Dict[str, bytes],
    insert_col_idx: int = 2,
    picture_header: str = "GOODS PICTURE",
    match_col_name: str = "รหัสสินค้า",
    min_row_height: float = 80,
    col_width: float = 18,
):
    """
    Inserts a new column (default column 2) into buyer file, and puts pictures there.
    Match uses buyer file column 'รหัสสินค้า' (default).
    """
    wb = openpyxl.load_workbook(buyer_xlsx_path)
    ws = wb.active

    header_row = find_header_row(ws)
    match_col = find_col_by_header(ws, header_row, match_col_name)

    if match_col is None:
        # fallback: try barcode
        match_col = find_col_by_header(ws, header_row, "barcode")

    if match_col is None:
        wb.save(buyer_xlsx_path)
        return

    # insert picture column
    ws.insert_cols(insert_col_idx, 1)
    ws.cell(header_row, insert_col_idx).value = picture_header
    ws.column_dimensions[openpyxl.utils.get_column_letter(insert_col_idx)].width = col_width

    # shift match_col if needed
    if match_col >= insert_col_idx:
        match_col += 1

    # loop rows
    for r in range(header_row + 1, ws.max_row + 1):
        key = norm_text(ws.cell(r, match_col).value)
        if not key:
            continue

        # match directly by code
        img_bytes = img_map.get(key)

        # fallback: remove spaces and compare
        if img_bytes is None:
            key2 = key.replace(" ", "")
            for k in img_map.keys():
                if k.replace(" ", "") == key2:
                    img_bytes = img_map[k]
                    break

        if img_bytes is None:
            continue

        # set row height
        if ws.row_dimensions[r].height is None or ws.row_dimensions[r].height < min_row_height:
            ws.row_dimensions[r].height = min_row_height

        # insert image
        img_obj = XLImage(BytesIO(img_bytes))
        img_obj.width = 90
        img_obj.height = 90
        ws.add_image(img_obj, ws.cell(r, insert_col_idx).coordinate)

    wb.save(buyer_xlsx_path)
